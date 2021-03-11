from openpyxl import Workbook


class Carro:
    """
    A classe carro tem como objetivo trazer dados necessários sobre o carro para que haja o dimensionamento do freio.
    Ela possibilita o uso desses dados para outras funcionalidades do carro.
    Para adicionar caracteristicas novas é preciso adicionar setters e properties, além de deixà-las
    privadas para q haja proteção desses atributos.
    """

    def __init__(self, nome, massa=1, dist_eixos=1,
                 alt_cg=1, dist_eix_dianteiro=1, dist_eixo_traseiro=1,
                 diam_cilindro_mestre=1, raio_pneu_dianteiro=1,
                 raio_pneu_traseiro=1, raio_disco_dianteiro=1, raio_disco_traseiro=1,
                 coef_atrito=1, relacao_pedal=1, diam_pinca_dianteira=1,
                 diam_pinca_traseira=1, coef_atrito_pinca=1, area_embolo=1,
                 numero_embolo=1, ):
        # Nome do carro
        self.__nome = nome.upper()

        # Massa(kg)
        self.__massa = massa

        # Distancia entra eixos(m)
        self.__dist_eixos = dist_eixos

        # ALtura do Centro de gravidade(m)
        self.__alt_cg = alt_cg

        # Distancia horizontal do centro de gravidade para o eixo dianteiro(m)
        self.__dist_eix_dianteiro = dist_eix_dianteiro

        # Distancia horizontal do centro de gravidade para o eixo  traseiro(m)
        self.__dist_eixo_traseiro = dist_eixo_traseiro

        # Diametro do cilindro mestre(m)
        self.__diam_cilindro_mestre = diam_cilindro_mestre

        # Raio do pneu dianteiro(m)
        self.__raio_pneu_dianteiro = raio_pneu_dianteiro

        # Raio do pneu dianteiro(m)
        self.__raio_pneu_traseiro = raio_pneu_traseiro

        # Raio do disco de freio do eixo dianteiro(m)
        self.__raio_disco_dianteiro = raio_disco_dianteiro

        # Raio do disco de  freio do eixo traseio(m)
        self.__raio_disco_traseiro = raio_disco_traseiro

        # coeficiente de atrito
        self.__coef_atrito = coef_atrito

        # Relação do pedal
        self.__relacao_pedal = relacao_pedal

        # Diametro da pinça traseira(m)
        self.__diam_pinca_traseira = diam_pinca_traseira

        # Diametro da pinça dianteira(m)
        self.__diam_pinca_dianteira = diam_pinca_dianteira

        # coeficiente de atrito dapinça com o disco
        self.__coef_atrito_pinca = coef_atrito_pinca

        # Área do embolo(m^2)
        self.__area_embolo = area_embolo

        # Numero de êmbolos
        self.__numero_embolo = numero_embolo

    def __str__(self):
        return f"""Dados do {self.nome}: 

massa: {self.massa}
distancia entre os eixos: {self.dist_eixos}
altura do centro de gavidande: {self.alt_cg}
Distancia horizontal do centro de gravidade para o eixo dianteiro: {self.dist_eix_dianteiro}
Distancia horizontal do centro de gravidade para o eixo traseiro: {self.dist_eixo_traseiro}
Diamentro do cilindro mestre: {self.diam_cilindro_mestre}
Raio do pneu dianteiro: {self.raio_pneu_dianteiro}
Raio do pneu traseiro: {self.raio_pneu_traseiro}
Raio do disco dianteiro: {self.raio_disco_dianteiro}
Raio do disco traseiro: {self.raio_disco_traseiro}
Coeficiente de atrito: {self.coef_atrito}
Relação do pedal: {self.relacao_pedal}
diametro da(s) pinça(s) dianteira(s): {self.diam_pinca_dianteira}
diametro da(s) pinça(s) traseiras(s): {self.diam_pinca_traseira} """

    # properties
    @property
    def nome(self):
        return self.__nome

    @property
    def massa(self):
        return f'{self.__massa} KG'

    @property
    def dist_eixos(self):
        return f'{self.__dist_eixos}m'

    @property
    def alt_cg(self):
        return f'{self.__alt_cg}m'

    @property
    def dist_eix_dianteiro(self):
        return f'{self.__dist_eix_dianteiro}m'

    @property
    def dist_eixo_traseiro(self):
        return f'{self.__dist_eixo_traseiro}m'

    @property
    def diam_cilindro_mestre(self):
        return f'{self.__diam_cilindro_mestre}m'

    @property
    def raio_pneu_dianteiro(self):
        return f'{self.__raio_pneu_dianteiro}m'

    @property
    def raio_pneu_traseiro(self):
        return f'{self.__raio_pneu_traseiro}m'

    @property
    def raio_disco_dianteiro(self):
        return f'{self.__raio_disco_dianteiro}m'

    @property
    def raio_disco_traseiro(self):
        return f'{self.__raio_disco_traseiro}m'

    @property
    def coef_atrito(self):
        return self.__coef_atrito

    @property
    def relacao_pedal(self):
        return self.__relacao_pedal

    @property
    def diam_pinca_dianteira(self):
        return f'{self.__diam_pinca_dianteira}m'

    @property
    def diam_pinca_traseira(self):
        return f'{self.__diam_pinca_traseira}m'

    @property
    def coef_atrito_pinca(self):
        return self.__coef_atrito_pinca

    @property
    def area_embolo(self):
        return f'{self.__area_embolo}m^2'

    @property
    def numero_embolo(self):
        return {self.__numero_embolo}

    # setters
    @nome.setter
    def nome(self, nome):
        self.__nome = nome

    @massa.setter
    def massa(self, massa):
        self.__massa = massa

    @dist_eixos.setter
    def dist_eixos(self, dist_eixos):
        self.__dist_eixos = dist_eixos

    @alt_cg.setter
    def alt_cg(self, alt_cg):
        self.__alt_cg = alt_cg

    @dist_eix_dianteiro.setter
    def dist_eix_dianteiro(self, dist_eix_dianteiro):
        self.__dist_eix_dianteiro = dist_eix_dianteiro

    @dist_eixo_traseiro.setter
    def dist_eixo_traseiro(self, dist_eixo_traseiro):
        self.__dist_eixo_traseiro = dist_eixo_traseiro

    @diam_cilindro_mestre.setter
    def diam_cilindro_mestre(self, diam_cilindro_mestre):
        self.__diam_cilindro_mestre = diam_cilindro_mestre

    @raio_pneu_dianteiro.setter
    def raio_pneu_dianteiro(self, raio_pneu_dianteiro):
        self.__raio_pneu_dianteiro = raio_pneu_dianteiro

    @raio_pneu_traseiro.setter
    def raio_pneu_traseiro(self, raio_pneu_traseiro):
        self.__raio_pneu_traseiro = raio_pneu_traseiro

    @raio_disco_dianteiro.setter
    def raio_disco_dianteiro(self, raio_disco_dianteiro):
        self.__raio_disco_dianteiro = raio_disco_dianteiro

    @raio_disco_traseiro.setter
    def raio_disco_traseiro(self, raio_disco_traseiro):
        self.__raio_disco_traseiro = raio_disco_traseiro

    @coef_atrito.setter
    def coef_atrito(self, coef_atrito):
        self.__coef_atrito = coef_atrito

    @relacao_pedal.setter
    def relacao_pedal(self, relacao_pedal):
        self.__relacao_pedal = relacao_pedal

    @diam_pinca_dianteira.setter
    def diam_pinca_dianteira(self, diam_pinca_dianteira):
        self.__diam_pinca_dianteira = diam_pinca_dianteira

    @diam_pinca_traseira.setter
    def diam_pinca_traseira(self, diam_pinca_traseira):
        self.__diam_pinca_traseira = diam_pinca_traseira

    @coef_atrito_pinca.setter
    def coef_atrito_pinca(self, coef_atrito_pinca):
        self.__coef_atrito_pinca = coef_atrito_pinca

    @area_embolo.setter
    def area_embolo(self, area_embolo):
        self.__area_embolo = area_embolo

    @numero_embolo.setter
    def numero_embolo(self, numero_embolo):
        self.__numero_embolo = numero_embolo

    # metodos estáticos
    @staticmethod
    def gravidade():
        return 9.81

    @staticmethod
    def precisao():
        return 4

    # metodos privados
    def __peso(self):
        peso = self.__massa * self.gravidade()
        return peso

    def __peso_estatico_diant(self):
        w = self.__peso() * self.__dist_eixo_traseiro / self.__dist_eixos
        return w.__round__(self.precisao())

    def __peso_estatico_traseiro(self):
        w = self.__peso() - self.__peso_estatico_diant()
        return w.__round__(self.precisao())

    def __aceleracao(self):
        a = self.gravidade() * self.__coef_atrito
        return - a.__round__(self.precisao())

    def __carga_din_dianteira(self):
        WF = self.__peso() * (self.__dist_eixo_traseiro / self.__dist_eixos
                              - (self.__aceleracao() / self.gravidade()) * self.__alt_cg / self.__dist_eixos)
        return WF.__round__(self.precisao())

    def __carga_din_traseira(self):
        WR = self.__peso() - self.__carga_din_dianteira()
        return WR.__round__(self.precisao())

    def __forca_transversal(self):
        fp = self.__carga_din_dianteira() * self.__coef_atrito
        return fp.__round__(self.precisao())

    def __momento_rodas_dianteiras(self):
        mf = self.__forca_transversal() * self.__raio_pneu_dianteiro

        return mf.__round__(self.precisao())

    def __forca_eixo_dianteiro(self):
        fd = self.__momento_rodas_dianteiras() / self.__raio_disco_dianteiro
        return fd.__round__(self.precisao())

    def __forca_disc(self):
        fd = self.__forca_eixo_dianteiro() / 2
        return fd

    def __forca_pastilhas(self):
        f = self.__forca_disc() * self.__coef_atrito_pinca
        return f.__round__(self.precisao())

    def __pressao(self):
        p = self.__forca_pastilhas() / (self.__area_embolo * self.__numero_embolo)
        return p.__round__(self.precisao())

    def __forca_cilindro(self):
        area_cilindro = 3.14159 * (self.__diam_cilindro_mestre / 2) ** 2
        fc = self.__pressao() * area_cilindro
        return fc

    def __forca_acionamento(self):
        fa = self.__forca_cilindro() / self.__relacao_pedal
        return fa

    @property
    def peso(self):
        return f'{self.__peso()} N'

    @property
    def peso_estatico_diant(self):
        return f'{self.__peso_estatico_diant()} N'

    @property
    def peso_estatico_traseiro(self):
        return f'{self.__peso_estatico_traseiro()} N'

    @property
    def aceleracao(self):
        return f'{self.__aceleracao()} m/s^2'

    @property
    def carga_din_traseira(self):
        return f'{self.__carga_din_traseira()} N'

    @property
    def carga_din_dianteira(self):
        return f'{self.__carga_din_dianteira()} N'

    @property
    def forca_transversal(self):
        return f'{self.__forca_transversal()} N * m'

    @property
    def momento_rodas_dianteiras(self):
        return f'{self.__momento_rodas_dianteiras()} N * M'

    @property
    def forca_eixo_dianteiro(self):
        return f'{self.__forca_eixo_dianteiro()} N'

    @property
    def forca_disc(self):
        return f'{self.__forca_disc()} N'

    @property
    def forca_pastilha(self):
        return f'{self.__forca_disc()} N'

    @property
    def pressao(self):
        return f'{self.__pressao()} N'

    @property
    def forca_cilindro(self):
        return f'{self.__forca_cilindro()} N'

    @property
    def forca_acionamento(self):
        return f'{self.__forca_acionamento()} N'

    @nome.setter
    def nome(self, value):
        self._nome = value

    def criar_planilha(self):
        wb = Workbook()
        ws = wb.active

        ws['A1'] = f'Dados do {self.nome}'
        ws['A2'] = 'Massa:'
        ws['A3'] = 'Distancia entre eixos'
        ws['A4'] = 'Altura do centro de gavidande:'
        ws['A5'] = 'Distancia horizontal do centro de gravidade para o eixo dianteiro:'
        ws['A6'] = 'Distancia horizontal do centro de gravidade para o eixo trsaseiro:'
        ws['A7'] = 'Diametro do cilindro mestre:'
        ws['A8'] = 'Raio do pneu dianteiro:'
        ws['A9'] = 'Raio do pneu traseiro:'
        ws['A10'] = 'Raio do disco dianteiro:'
        ws['A11'] = 'Raio do disco traseiro:'
        ws['A12'] = 'Coeficiente de atrito:'
        ws['A13'] = 'Relação do pedal:'
        ws['A14'] = 'diametro da(s) pinça(s) dianteira(s):'
        ws['A15'] = 'diametro da(s) pinça(s) traseiras(s):'

        ws['B2'] = f'{self.massa}'
        ws['B3'] = f'{self.dist_eixos}'
        ws['B4'] = f'{self.alt_cg}'
        ws['B5'] = f'{self.dist_eix_dianteiro}'
        ws['B6'] = f'{self.dist_eixo_traseiro}'
        ws['B7'] = f'{self.diam_cilindro_mestre}'
        ws['B8'] = f'{self.raio_pneu_dianteiro}'
        ws['B9'] = f'{self.raio_pneu_traseiro}'
        ws['B10'] = f'{self.raio_disco_dianteiro}'
        ws['B11'] = f'{self.raio_disco_traseiro}'
        ws['B12'] = f'{self.coef_atrito}'
        ws['B13'] = f'{self.relacao_pedal}'
        ws['B14'] = f'{self.diam_pinca_dianteira}'
        ws['B15'] = f'{self.diam_pinca_traseira}'
        wb.save()

'''        def __peso(self):
            peso = self.__massa * self.gravidade()
            return peso

        def __peso_estatico_diant(self):
            w = self.__peso() * self.__dist_eixo_traseiro / self.__dist_eixos
            return w.__round__(self.precisao())

        def __peso_estatico_traseiro(self):
            w = self.__peso() - self.__peso_estatico_diant()
            return w.__round__(self.precisao())

        def __aceleracao(self):
            a = self.gravidade() * self.__coef_atrito
            return - a.__round__(self.precisao())

        def __carga_din_dianteira(self):
            WF = self.__peso() * (self.__dist_eixo_traseiro / self.__dist_eixos
                                  - (self.__aceleracao() / self.gravidade()) * self.__alt_cg / self.__dist_eixos)
            return WF.__round__(self.precisao())

        def __carga_din_traseira(self):
            WR = self.__peso() - self.__carga_din_dianteira()
            return WR.__round__(self.precisao())

        def __forca_transversal(self):
            fp = self.__carga_din_dianteira() * self.__coef_atrito
            return fp.__round__(self.precisao())

        def __momento_rodas_dianteiras(self):
            mf = self.__forca_transversal() * self.__raio_pneu_dianteiro

            return mf.__round__(self.precisao())

        def __forca_eixo_dianteiro(self):
            fd = self.__momento_rodas_dianteiras() / self.__raio_disco_dianteiro
            return fd.__round__(self.precisao())

        def __forca_disc(self):
            fd = self.__forca_eixo_dianteiro() / 2
            return fd

        def __forca_pastilhas(self):
            f = self.__forca_disc() * self.__coef_atrito_pinca
            return f.__round__(self.precisao())

        def __pressao(self):
            p = self.__forca_pastilhas() / (self.__area_embolo * self.__numero_embolo)
            return p.__round__(self.precisao())

        def __forca_cilindro(self):
            area_cilindro = 3.14159 * (self.__diam_cilindro_mestre / 2) ** 2
            fc = self.__pressao() * area_cilindro
            return fc

        def __forca_acionamento(self):
            fa = self.__forca_cilindro() / self.__relacao_pedal
            return fa
'''




class Arquivo:
    def __init__(self, nome):
        if nome.find('.txt') == -1:
            self.nome = nome + '.txt'
        else:
            self.nome = nome
        if not self.arquivo_existe():
            self.criar_arquivo()

    def criar_arquivo(self):
        try:
            a = open(self.nome, 'wt+')
            a.close()
        except:
            print('houve um ERR0')
        else:
            print(f'{self.nome} foi criado')

    def arquivo_existe(self):
        try:
            a = open(self.nome, 'rt')
            a.close()
        except FileNotFoundError:
            return False
        else:
            return True


class Arquivo_baja(Arquivo):
    def __init__(self, nome):
        super().__init__(nome)

    def adicionar_carro(self, carro):
        try:
            a = open(self.nome, 'at')
        except:
            print('houve um erro no arquivo!!!')
        else:
            try:
                a.write(f'{carro};\n')
                a.close()
            except:
                print("Naõ foi possível adicionar Carro")

        pass

    def selecionar_carro(self):
        n = 1
        a = (open(self.nome, 'rt'))
        for linha in a:
            if linha.find("Dados") != -1:
                print(f"{n}-{linha[9:]}")
                n += 1
        opcao = input('qual carro deseja escolher:')
        return opcao

    def alterar_carro(self):
        pass

    def excluir_carro(self):
        pass

    def criar_planilha(self):
        pass

    def criar_relatorio(self):
        pass


